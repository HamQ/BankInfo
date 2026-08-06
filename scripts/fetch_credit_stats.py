# -*- coding: utf-8 -*-
"""
信用卡重要业务及财务资讯揭露 — 解析与入库
数据源: https://www.banking.gov.tw/ch/home.jsp?id=591&...&dataserno=21207
运行: python fetch_credit_stats.py
需要环境变量: SUPABASE_URL, SUPABASE_SERVICE_KEY
"""

import os, sys, re, io, zipfile, tempfile
from datetime import date
from urllib.request import Request, urlopen

import openpyxl
from supabase import create_client, Client

# ─── 配置 ───────────────────────────────────────────
FSC_PAGE_URL = "https://www.banking.gov.tw/ch/home.jsp?id=591&parentpath=0,590&mcustomize=multimessage_view.jsp&dataserno=21207&dtable=Disclosure"
USER_AGENT = "Mozilla/5.0 (compatible; BankInfoBot/1.0)"

# ─── 银行名称映射 ────────────────────────────────────
# Excel 中的名称可能带有全角空格，此处做清洗映射
BANK_NAME_MAP = {
    "臺灣銀行": "004",
    "臺灣土地銀行": "005",
    "合作金庫商業銀行": "006",
    "第一商業銀行": "007",
    "華南商業銀行": "008",
    "彰化商業銀行": "009",
    "上海商業儲蓄銀行": "011",
    "台北富邦商業銀行": "012",
    "國泰世華商業銀行": "013",
    "高雄銀行": "016",
    "兆豐國際商業銀行": "017",
    "花旗(台灣)商業銀行": "021",
    "臺灣中小企業銀行": "050",
    "渣打國際商業銀行": "052",
    "台中商業銀行": "053",
    "滙豐(台灣)商業銀行": "081",
    "華泰商業銀行": "101",
    "臺灣新光商業銀行": "102",
    "陽信商業銀行": "103",
    "三信商業銀行": "108",
    "聯邦商業銀行": "803",
    "遠東國際商業銀行": "805",
    "元大商業銀行": "806",
    "永豐商業銀行": "807",
    "玉山商業銀行": "808",
    "凱基商業銀行": "809",
    "星展(台灣)商業銀行": "810",
    "台新國際商業銀行": "812",
    "安泰商業銀行": "815",
    "中國信託商業銀行": "822",
    "台灣樂天信用卡股份有限公司": "RC001",
    "台灣美國運通國際(股)公司": "AE001",
}

def fetch_page(url: str) -> str:
    """抓取页面 HTML（忽略 SSL 验证以兼容金管局旧证书）"""
    import ssl
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(req, context=ctx, timeout=30) as resp:
        return resp.read().decode("utf-8")

def find_zip_url(html: str) -> str:
    """从页面中找到最新的 ZIP 下载链接"""
    match = re.search(r'https?://[^"\s]+\.zip', html)
    if not match:
        raise RuntimeError("页面中未找到 ZIP 下载链接")
    return match.group(0)

def download_zip(zip_url: str) -> bytes:
    """下载 ZIP 文件内容"""
    import ssl
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = Request(zip_url, headers={"User-Agent": USER_AGENT})
    with urlopen(req, context=ctx, timeout=60) as resp:
        return resp.read()

def parse_excel(zip_bytes: bytes) -> tuple:
    """解析 ZIP 中的 Excel 文件，返回 (report_month, rows)"""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        xlsx_files = [n for n in zf.namelist() if n.endswith('.xlsx')]
        if not xlsx_files:
            raise RuntimeError("ZIP 中未找到 .xlsx 文件")
        with zf.open(xlsx_files[0]) as f:
            wb = openpyxl.load_workbook(f, data_only=True)

    ws = wb[wb.sheetnames[0]]

    # 提取报表月份 — 通常在第三行的第 10 列附近
    report_month = None
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=14, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                m = re.search(r'(\d{2,3})\s*年\s*(\d{1,2})\s*月', str(cell))
                if m:
                    year = int(m.group(1)) + 1911  # 民国年 → 西元年
                    month = int(m.group(2))
                    report_month = date(year, month, 1)
                    break
        if report_month:
            break

    if not report_month:
        raise RuntimeError("无法从 Excel 中解析报表月份")

    # 解析数据行（从第 4 行表头之后开始）
    rows = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=14, values_only=True):
        bank_name_raw = row[0]
        if not bank_name_raw or not isinstance(bank_name_raw, str):
            continue
        bank_name = re.sub(r'[\s\u3000]+', '', str(bank_name_raw).strip())

        if bank_name not in BANK_NAME_MAP:
            print(f"  ⚠ 未知银行: {bank_name_raw}")
            continue

        def to_int(v):
            try:
                if v is None: return None
                return int(float(v))
            except (ValueError, TypeError):
                return None

        def to_num(v):
            try:
                if v is None: return None
                return round(float(v), 2)
            except (ValueError, TypeError):
                return None

        rows.append({
            "code": BANK_NAME_MAP[bank_name],
            "cards_in_circulation":     to_int(row[1]),
            "active_cards":             to_int(row[2]),
            "cards_issued_month":       to_int(row[3]),
            "cards_stopped_month":      to_int(row[4]),
            "revolving_balance":        to_int(row[5]),
            "installment_balance":      to_int(row[6]),
            "transaction_volume":       to_int(row[7]),
            "cash_advance_volume":      to_int(row[8]),
            "delinquency_3m_ratio":     to_num(row[9]),
            "delinquency_6m_ratio":     to_num(row[10]),
            "bad_debt_coverage_ratio":  to_num(row[11]),
            "bad_debt_writeoff_month":  to_int(row[12]),
            "bad_debt_writeoff_ytd":    to_int(row[13]),
        })

    return report_month, rows

def upsert_to_supabase(supabase: Client, report_month: date, rows: list, source_url: str):
    """将解析结果 UPSERT 到 Supabase"""
    # 获取 bank code → id 映射
    resp = supabase.table("banks").select("id,code").execute()
    code_to_id = {b["code"]: b["id"] for b in resp.data}

    upsert_count = 0
    for row in rows:
        bank_id = code_to_id.get(row["code"])
        if not bank_id:
            print(f"  ⚠ code={row['code']} 在 banks 表中不存在")
            continue

        record = {
            "bank_id": bank_id,
            "report_month": report_month.isoformat(),
            "cards_in_circulation": row["cards_in_circulation"],
            "active_cards": row["active_cards"],
            "cards_issued_month": row["cards_issued_month"],
            "cards_stopped_month": row["cards_stopped_month"],
            "revolving_balance": row["revolving_balance"],
            "installment_balance": row["installment_balance"],
            "transaction_volume": row["transaction_volume"],
            "cash_advance_volume": row["cash_advance_volume"],
            "delinquency_3m_ratio": row["delinquency_3m_ratio"],
            "delinquency_6m_ratio": row["delinquency_6m_ratio"],
            "bad_debt_coverage_ratio": row["bad_debt_coverage_ratio"],
            "bad_debt_writeoff_month": row["bad_debt_writeoff_month"],
            "bad_debt_writeoff_ytd": row["bad_debt_writeoff_ytd"],
            "source_url": source_url,
        }

        supabase.table("monthly_credit_stats") \
            .upsert(record, on_conflict="bank_id,report_month") \
            .execute()
        upsert_count += 1

    print(f"  ✅ UPSERT {upsert_count} 笔记录")

def main():
    print("🔍 台湾信用卡情报雷达 — 信用卡月报同步")
    print(f"📄 数据源: {FSC_PAGE_URL}")

    # 1. 抓取页面，定位 ZIP
    print("\n[1/4] 抓取金管局页面...")
    html = fetch_page(FSC_PAGE_URL)
    zip_url = find_zip_url(html)
    print(f"  ✅ ZIP 链接: {zip_url}")

    # 2. 下载 ZIP
    print("[2/4] 下载 ZIP 文件...")
    zip_bytes = download_zip(zip_url)
    print(f"  ✅ 下载完成 ({len(zip_bytes):,} bytes)")

    # 3. 解析 Excel
    print("[3/4] 解析 Excel 数据...")
    report_month, rows = parse_excel(zip_bytes)
    print(f"  ✅ 报表月份: {report_month}, 解析 {len(rows)} 家银行")

    # 4. 入库 Supabase
    print("[4/4] 写入 Supabase...")
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        print("❌ 请设置 SUPABASE_URL 和 SUPABASE_SERVICE_KEY")
        sys.exit(1)
    supabase = create_client(url, key)
    upsert_to_supabase(supabase, report_month, rows, zip_url)

    print("\n🎉 信用卡月报同步完成！")

if __name__ == "__main__":
    main()
