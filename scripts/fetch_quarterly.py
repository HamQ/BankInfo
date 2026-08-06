# -*- coding: utf-8 -*-
"""
季度补充数据同步 — 数位存款 + 逾放资料 (v5)
NPL 解析器重写：适配实际 Excel 结构（原始财务数据 → 计算比率）
"""
import os, sys, re, io, zipfile, ssl
from datetime import date
from urllib.request import Request, urlopen
from urllib.parse import quote

import openpyxl
from supabase import create_client

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SOURCES = {
    "digital_acct": {
        "url": "https://www.banking.gov.tw/ch/home.jsp?id=591&parentpath=0,590&mcustomize=multimessage_view.jsp&dataserno=201911270001&dtable=Disclosure",
        "table": "quarterly_digital_acct_stats",
        "desc": "数位存款帐户",
        "min_year": 113,
        "has_unique": True,
    },
    "npl": {
        "url": "https://www.banking.gov.tw/ch/home.jsp?id=591&parentpath=0,590&mcustomize=multimessage_view.jsp&dataserno=201202130001&dtable=Disclosure",
        "table": "quarterly_npl_stats",
        "desc": "逾放资料",
        "min_year": 110,
        "has_unique": True,
    },
}

USER_AGENT = "Mozilla/5.0 (compatible; BankInfoBot/1.0)"

def ssl_ctx():
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx

def fetch(url):
    req = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(req, context=ssl_ctx(), timeout=30) as r:
        return r.read().decode("utf-8")

def find_zips_digital(html, min_year):
    pattern = r'(https?://[^"\s]*?\.zip)'
    urls = re.findall(pattern, html)
    results = []
    for u in urls:
        m = re.search(r'(\d{2,3})Q(\d)', u)
        if m:
            y, q = int(m.group(1)), int(m.group(2))
            if y < min_year or q < 1 or q > 4:
                continue
            month = (q - 1) * 3 + 1
            results.append((date(y + 1911, month, 1), u))
    return sorted(results, reverse=True)

def find_zips_npl(html, min_year):
    pattern = r'(https?://[^"\s]*?\.zip)'
    urls = re.findall(pattern, html)
    results = []
    seen = set()
    for u in urls:
        y, mo = None, None
        
        m = re.search(r'(\d{2,3})[_-](\d{2})', u)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
        
        if y is None:
            m = re.search(r'(\d{3})(\d{2})(?:\(\d+\))?\.zip', u)
            if m:
                y, mo = int(m.group(1)), int(m.group(2))
            else:
                m = re.search(r'(\d{2})(\d{2})(?:\(\d+\))?\.zip', u)
                if m:
                    y2, mo2 = int(m.group(1)), int(m.group(2))
                    if y2 >= 90:
                        y, mo = y2, mo2
        
        if y is None or mo is None:
            continue
        if y < min_year or mo < 1 or mo > 12:
            continue
        
        rq = date(y + 1911, mo, 1)
        key = rq.isoformat()
        if key not in seen:
            seen.add(key)
            results.append((rq, u))
    
    return sorted(results, reverse=True)

def download_zip(url):
    safe = re.sub(r"[^\x00-\x7F]+", lambda m: quote(m.group(0), encoding="utf-8"), url)
    req = Request(safe, headers={"User-Agent": USER_AGENT})
    with urlopen(req, context=ssl_ctx(), timeout=60) as r:
        return r.read()

def read_excel_from_zip(zip_bytes):
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        
        for n in names:
            if n.endswith(".xlsx"):
                with zf.open(n) as f:
                    return openpyxl.load_workbook(f, data_only=True), "xlsx"
        
        for n in names:
            if n.endswith(".xls"):
                import xlrd
                with zf.open(n) as f:
                    return xlrd.open_workbook(file_contents=f.read()), "xls"
        
        for n in names:
            if n.endswith(".ods"):
                import pandas as pd
                with zf.open(n) as f:
                    return pd.read_excel(io.BytesIO(f.read()), engine="odf"), "ods"
        
        return None, "ZIP 内无 Excel 文件: " + ", ".join(names[:5])

def iter_excel_rows(wb, fmt):
    if fmt == "xlsx":
        ws = wb[wb.sheetnames[0]]
        return [[c for c in row] for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]
    elif fmt == "xls":
        sh = wb.sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    else:
        df = wb
        return [list(df.columns)] + df.values.tolist()

def find_report_date(rows):
    for row in rows[:8]:
        for c in row:
            if c and isinstance(c, str):
                m = re.search(r'(\d{2,3})\s*年\s*(\d{1,2})\s*月', str(c))
                if m:
                    y, mo = int(m.group(1)), int(m.group(2))
                    return date(y + 1911, mo, 1)
    return None

def parse_digital_acct(zip_bytes, fallback_url):
    """解析数位存款 Excel — 列A是银行全名"""
    wb, fmt = read_excel_from_zip(zip_bytes)
    if wb is None:
        return None, [], fmt
    
    rows = iter_excel_rows(wb, fmt)
    rq = find_report_date(rows)
    if not rq:
        return None, [], "未找到报告月份"
    
    def ti(v):
        if v is None: return None
        try: return int(float(str(v).replace(",", "")))
        except: return None
    
    SKIP_KW = ["数位", "數位", "中华民国", "中華民國", "单位:", "單位:", "第一类", "第一類", "第二类", "第二類", "统计", "統計"]
    
    data_rows = []
    for row in rows:
        if len(row) < 2:
            continue
        col0 = row[0]
        if col0 is None:
            continue
        col0_str = str(col0).strip()
        
        if not col0_str:
            continue
        if any(kw in col0_str for kw in SKIP_KW):
            continue
        if "总计" in col0_str or "總計" in col0_str or "合计" in col0_str:
            continue
        if col0_str.isdigit():
            continue
        
        data_rows.append({
            "name": col0_str,
            "type1_accounts": ti(row[1]),
            "type2_accounts": ti(row[2]),
            "type3_accounts": ti(row[3]) if len(row) > 3 else None,
            "total_accounts": ti(row[4]) if len(row) > 4 else None,
        })
    
    return rq, data_rows, None

def parse_npl(zip_bytes, fallback_url):
    """解析逾放资料 Excel — 实际结构:
    银行别 | 存款 | 税前盈余(累计) | 放款总额 | 逾期放款总额 | 备抵呆帐 | ...
    逾放比率 = 逾期放款 / 放款总额 * 100
    备抵覆盖率 = 备抵呆帐 / 逾期放款 * 100
    单位: 百万元
    """
    wb, fmt = read_excel_from_zip(zip_bytes)
    if wb is None:
        return None, [], fmt
    
    rows = iter_excel_rows(wb, fmt)
    rq = find_report_date(rows)
    if not rq:
        return None, [], "未找到报告月份"
    
    def ti(v):
        if v is None: return None
        try: return int(float(str(v).replace(",", "")))
        except: return None
    
    SKIP_KW = ["总计", "總計", "合计", "银行别", "銀行別", "存款", "税前", "稅前", "放款", "逾期", "备抵", "備抵",
                "统计", "統計", "月份", "单位:", "單位:", "申报", "申報"]
    
    data_rows = []
    for row in rows:
        if len(row) < 6:
            continue
        col0 = str(row[0]).strip() if row[0] is not None else ""
        if not col0 or len(col0) < 2:
            continue
        if any(kw in col0 for kw in SKIP_KW):
            continue
        # 银行名至少包含一个银行相关汉字
        if not any(c in col0 for c in ["银行", "銀行", "金库", "金庫", "储蓄", "儲蓄", "信用", "商业", "商業", "農會", "渔會", "漁會"]):
            continue
        
        total_loans = ti(row[3])       # 放款总额
        overdue_loans = ti(row[4])     # 逾期放款总额
        loan_allowance = ti(row[5])    # 备抵呆帐
        
        # 计算比率 (百分比)
        npl_ratio = None
        if total_loans and overdue_loans and total_loans > 0:
            npl_ratio = round(overdue_loans / total_loans * 100, 2)
        
        coverage_ratio = None
        if overdue_loans and loan_allowance and overdue_loans > 0:
            coverage_ratio = round(loan_allowance / overdue_loans * 100, 2)
        
        data_rows.append({
            "name": col0,
            "deposits": ti(row[1]),
            "pre_tax_profit": ti(row[2]),
            "total_loans": total_loans,
            "overdue_loans": overdue_loans,
            "loan_allowance": loan_allowance,
            "npl_ratio": npl_ratio,
            "coverage_ratio": coverage_ratio,
        })
    
    return rq, data_rows, None

def run_source(supabase, cfg, name_to_id):
    table = cfg["table"]
    desc = cfg["desc"]
    min_year = cfg.get("min_year", 113)
    has_unique = cfg.get("has_unique", True)
    
    # 获取已有的 (bank_id, report_quarter) 组合
    existing_pairs = set()
    for r in supabase.table(table).select("bank_id,report_quarter").execute().data:
        existing_pairs.add((r["bank_id"], r["report_quarter"]))
    
    print(f"\n{'='*50}")
    print(f"{desc} — 现有 {len(existing_pairs)} 条记录")
    
    html = fetch(cfg["url"])
    
    if "npl" in table:
        months = find_zips_npl(html, min_year)
    else:
        months = find_zips_digital(html, min_year)
    print(f"发现 {len(months)} 个季度")
    
    new_count = 0
    for rq, zurl in months:
        roc_y = rq.year - 1911
        roc_m = rq.month
        rq_str = rq.isoformat()
        
        print(f"  新季度: 民国{roc_y}年{roc_m}月")
        try:
            zb = download_zip(zurl)
            if "digital" in table:
                pq, data_rows, err = parse_digital_acct(zb, zurl)
            else:
                pq, data_rows, err = parse_npl(zb, zurl)
            
            if pq is None:
                print(f"    跳过: {err}")
                continue
            
            pq_str = pq.isoformat()
            inserted = 0
            skipped_dup = 0
            for r in data_rows:
                bid = name_to_id.get(r["name"])
                if not bid:
                    if inserted < 3:
                        print(f"    未匹配银行: {r['name']}")
                    continue
                
                pair = (bid, pq_str)
                if pair in existing_pairs:
                    skipped_dup += 1
                    continue
                
                record = {
                    "bank_id": bid,
                    "report_quarter": pq_str,
                    "source_url": zurl.replace(".zip", ".pdf"),
                }
                if "digital" in table:
                    record.update({
                        "type1_accounts": r.get("type1_accounts"),
                        "type2_accounts": r.get("type2_accounts"),
                        "type3_accounts": r.get("type3_accounts"),
                        "total_accounts": r.get("total_accounts"),
                    })
                else:
                    record.update({
                        "bank_type": r.get("bank_type"),
                        "deposits": r.get("deposits"),
                        "pre_tax_profit": r.get("pre_tax_profit"),
                        "total_loans": r.get("total_loans"),
                        "overdue_loans": r.get("overdue_loans"),
                        "loan_allowance": r.get("loan_allowance"),
                        "npl_ratio": r.get("npl_ratio"),
                        "coverage_ratio": r.get("coverage_ratio"),
                    })
                
                try:
                    if has_unique:
                        supabase.table(table).upsert(record, on_conflict="bank_id,report_quarter").execute()
                    else:
                        supabase.table(table).insert(record).execute()
                    existing_pairs.add(pair)
                    inserted += 1
                except Exception as e:
                    print(f"    单条失败 [{r['name']}]: {e}")
            
            if skipped_dup > 0:
                print(f"    跳过 {skipped_dup} 条重复")
            print(f"    入库 {inserted} 笔")
            new_count += inserted
        except Exception as e:
            import traceback
            print(f"    失败: {e}")
            traceback.print_exc()
    
    print(f"{desc} 完成，新增 {new_count} 笔")

def main():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        print("请设置 SUPABASE_URL 和 SUPABASE_SERVICE_KEY")
        sys.exit(1)
    supabase = create_client(url, key)
    
    resp = supabase.table("banks").select("id,code,name,short_name").execute()
    name_to_id = {}
    for b in resp.data:
        name_to_id[b["name"]] = b["id"]
        if b.get("short_name"):
            name_to_id[b["short_name"]] = b["id"]
    print(f"已加载 {len(name_to_id)} 家银行映射")
    
    for name, cfg in SOURCES.items():
        try:
            run_source(supabase, cfg, name_to_id)
        except Exception as e:
            import traceback
            print(f"{cfg['desc']} 整体失败: {e}")
            traceback.print_exc()

if __name__ == "__main__":
    main()
