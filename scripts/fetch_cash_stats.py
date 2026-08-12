# -*- coding: utf-8 -*-
"""
现金卡重要业务及财务资讯揭露 — 解析与入库 (v2)
- 抓取页面全部月份 ZIP，只更新缺失月份
- URL 自动转义，同时显示 民国/公元 年份
"""
import os, sys, re, io, zipfile, ssl
from datetime import date
from urllib.request import Request, urlopen
from urllib.parse import quote

import openpyxl
from supabase import create_client, Client

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

FSC_PAGE_URL = "https://www.banking.gov.tw/ch/home.jsp?id=591&parentpath=0,590&mcustomize=multimessage_view.jsp&dataserno=21206&dtable=Disclosure"
USER_AGENT = "Mozilla/5.0 (compatible; BankInfoBot/1.0)"
MIN_ROC_YEAR = 113

BANK_NAME_MAP = {
    "第一商業銀行": "007", "華南商業銀行": "008", "台中商業銀行": "053",
    "臺灣銀行": "004", "臺灣土地銀行": "005", "合作金庫商業銀行": "006",
    "彰化商業銀行": "009", "上海商業儲蓄銀行": "011", "台北富邦商業銀行": "012",
    "國泰世華商業銀行": "013", "高雄銀行": "016", "兆豐國際商業銀行": "017",
    "花旗(台灣)商業銀行": "021", "臺灣中小企業銀行": "050", "渣打國際商業銀行": "052",
    "滙豐(台灣)商業銀行": "081", "華泰商業銀行": "101", "臺灣新光商業銀行": "102",
    "陽信商業銀行": "103", "三信商業銀行": "108", "聯邦商業銀行": "803",
    "遠東國際商業銀行": "805", "元大商業銀行": "806", "永豐商業銀行": "807",
    "玉山商業銀行": "808", "凱基商業銀行": "809", "星展(台灣)商業銀行": "810",
    "台新國際商業銀行": "812", "安泰商業銀行": "815", "中國信託商業銀行": "822",
}

def _ssl():
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx

def fetch_page(url):
    req = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(req, context=_ssl(), timeout=30) as r:
        return r.read().decode("utf-8")

def roc_label(y, m):
    return f"民国{y}年{m}月 (公元{y+1911}-{m:02d})"

def find_all_zips(html):
    pattern = r'(https?://[^\"\s]*?/(\d{3})(\d{2})_[^\"\s]*?\.zip)'
    matches = re.findall(pattern, html)
    month_map = {}
    for full, roc_y, roc_m in matches:
        y, m = int(roc_y), int(roc_m)
        if y < MIN_ROC_YEAR: continue
        month_map[(y, m)] = full
    return [(date(y+1911, m, 1), url) for (y, m), url in sorted(month_map.items(), reverse=True)]

def get_existing(supabase):
    return {r["report_month"] for r in supabase.table("monthly_cash_stats").select("report_month").execute().data}

def download_zip(zip_url):
    safe = re.sub(r"[^\x00-\x7F]+", lambda m: quote(m.group(0), encoding="utf-8"), zip_url)
    req = Request(safe, headers={"User-Agent": USER_AGENT})
    with urlopen(req, context=_ssl(), timeout=60) as r:
        return r.read()

def parse_excel(zip_bytes):
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        xlsx = [n for n in zf.namelist() if n.endswith(".xlsx")]
        if not xlsx: return None, []
        with zf.open(xlsx[0]) as f:
            wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rm = None
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=10, values_only=True):
        for c in row:
            if c and isinstance(c, str):
                m = re.search(r"(\d{2,3})\s*年\s*(\d{1,2})\s*月", str(c))
                if m: rm = date(int(m.group(1))+1911, int(m.group(2)), 1); break
        if rm: break
    if not rm: return None, []

    def ti(v):
        try: return None if v is None else int(float(v))
        except: return None
    def tn(v):
        try: return None if v is None else round(float(v), 2)
        except: return None

    rows = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, max_col=10, values_only=True):
        bn = row[0]
        if not bn or not isinstance(bn, str): continue
        bn = re.sub(r"[\s\u3000]+", "", str(bn).strip())
        if bn not in BANK_NAME_MAP: continue
        rows.append({"code": BANK_NAME_MAP[bn],
            "drawn_cards": ti(row[1]), "undrawn_cards": ti(row[2]),
            "contract_limit": ti(row[3]), "available_limit": ti(row[4]),
            "loan_balance": ti(row[5]), "delinquency_ratio": tn(row[6]),
            "provision_balance": ti(row[7]), "writeoff_month": ti(row[8]),
            "writeoff_ytd": ti(row[9])})
    return rm, rows

def upsert(supabase, rm, rows, src, c2i):
    n = 0
    for r in rows:
        bid = c2i.get(r["code"])
        if not bid: continue
        supabase.table("monthly_cash_stats").upsert({
            "bank_id": bid, "report_month": rm.isoformat(),
            "drawn_cards": r["drawn_cards"], "undrawn_cards": r["undrawn_cards"],
            "contract_limit": r["contract_limit"], "available_limit": r["available_limit"],
            "loan_balance": r["loan_balance"], "delinquency_ratio": r["delinquency_ratio"],
            "provision_balance": r["provision_balance"], "writeoff_month": r["writeoff_month"],
            "writeoff_ytd": r["writeoff_ytd"], "source_url": src.replace('.zip', '.pdf'),
        }, on_conflict="bank_id,report_month").execute()
        n += 1
    return n

def main():
    print("🏦 台湾银行卡片信息 - 现金卡月报同步 v2")
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        print("请设置 SUPABASE_URL 和 SUPABASE_SERVICE_KEY"); sys.exit(1)
    supabase = create_client(url, key)

    existing = get_existing(supabase)
    print(f"数据库已有 {len(existing)} 个月份")
    print("抓取金管局页面...")
    html = fetch_page(FSC_PAGE_URL)
    months = find_all_zips(html)
    print(f"页面中 {len(months)} 个可用月份 (自 {MIN_ROC_YEAR} 年起)")
    c2i = {b["code"]: b["id"] for b in supabase.table("banks").select("id,code").execute().data}

    new = 0
    for rm, zurl in months:
        ms = rm.isoformat()
        if ms in existing: continue
        ry, rm_val = rm.year - 1911, rm.month
        print(f"\n新月份: {roc_label(ry, rm_val)}")
        try:
            zb = download_zip(zurl)
            pm, rows = parse_excel(zb)
            if pm is None: print("  无法解析"); continue
            n = upsert(supabase, pm, rows, zurl, c2i)
            print(f"  入库 {n} 笔 ({len(rows)} 家机构)")
            new += n; existing.add(ms)
        except Exception as e:
            print(f"  失败: {e}")

    print("\n所有月份均为最新" if new == 0 else f"\n完成! 新增 {new} 笔记录")

if __name__ == "__main__":
    main()
